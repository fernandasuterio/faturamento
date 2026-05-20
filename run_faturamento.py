from google.cloud import bigquery
import pandas as pd
import os
import datetime
import warnings
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

client = bigquery.Client(project='meli-bi-data')

# ============================================================
# Query Templates  (placeholder {filtro_tipo} / {filtro_tipo_rv})
# ============================================================
QUERY_TEMPLATE = """
WITH TopPeriods AS (
  SELECT DISTINCT JSON_EXTRACT_SCALAR(SHP_SRM_PRIVN_PERIOD, '$.name') AS Name
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  ORDER BY Name DESC
  LIMIT 5
),
LatestVersion AS (
  SELECT SHP_SRM_PRIVN_TRANSACTION_ID, MAX(SHP_SRM_PRIVN_VERSION_ID) AS MaxVersion
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  GROUP BY SHP_SRM_PRIVN_TRANSACTION_ID
),
TraceDedup AS (
  SELECT PREINVOICE_ID, INVOICE_IDENTIFIER_STATUS
  FROM (
    SELECT PREINVOICE_ID, INVOICE_IDENTIFIER_STATUS,
           ROW_NUMBER() OVER (PARTITION BY PREINVOICE_ID ORDER BY INVOICE_LAST_UPDATE DESC) AS rn
    FROM WHOWNER.BT_SRM_INVOICE_FILE_TRACEABILITY
  )
  WHERE rn = 1
),
Base AS (
  SELECT
    C.SHP_SRM_PRIVN_TRANSACTION_ID,
    JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name') AS Name,
    C.SHP_SRM_PRIVN_COST_AMT,
    C.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
    C.SHP_SRM_PRIVN_PRE_INVOICE_ID,
    C.SHP_SRM_PRIVN_PAYLOAD,
    PRV.TYPE AS PROVIDER_TYPE,
    CASE
      WHEN PRV.TYPE = 'TAC-CPF' AND T.INVOICE_IDENTIFIER_STATUS IS NULL THEN 'nao_emite_nfs'
      WHEN T.INVOICE_IDENTIFIER_STATUS IN ('sap_status_pago_realizado', 'sap_status_pagado') THEN 'sap_status_pago_realizado'
      ELSE T.INVOICE_IDENTIFIER_STATUS
    END AS INVOICE_IDENTIFIER_STATUS
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES C
  JOIN LatestVersion LV
    ON C.SHP_SRM_PRIVN_TRANSACTION_ID = LV.SHP_SRM_PRIVN_TRANSACTION_ID
   AND C.SHP_SRM_PRIVN_VERSION_ID = LV.MaxVersion
  JOIN TopPeriods TP
    ON JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name') = TP.Name
  JOIN WHOWNER.BT_SHP_SRM_PROVIDERS PRV
    ON C.SHP_SRM_PRIVN_PROVIDER_ID = PRV.ID
  LEFT JOIN TraceDedup T
    ON CAST(C.SHP_SRM_PRIVN_PRE_INVOICE_ID AS STRING) = T.PREINVOICE_ID
  WHERE C.SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
    AND PRV.TYPE IN ('TAC-CNPJ', 'TAC-CPF', 'ETC', 'MEI')
),
PayloadKeyVals AS (
  SELECT
    b.SHP_SRM_PRIVN_TRANSACTION_ID,
    MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'external_route_id', JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS external_route_id
  FROM Base b,
       UNNEST(JSON_EXTRACT_ARRAY(b.SHP_SRM_PRIVN_PAYLOAD, '$.payload')) AS p
  GROUP BY b.SHP_SRM_PRIVN_TRANSACTION_ID
)
SELECT
  b.Name,
  b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
  b.INVOICE_IDENTIFIER_STATUS,
  COUNT(DISTINCT p.external_route_id)            AS qtd_rotas_distintas,
  COUNT(DISTINCT b.SHP_SRM_PRIVN_PRE_INVOICE_ID) AS qtd_pre_invoice_id_distintos,
  SUM(b.SHP_SRM_PRIVN_COST_AMT)                  AS SHP_SRM_PRIVN_COST_AMT
FROM Base b
LEFT JOIN PayloadKeyVals p
  ON b.SHP_SRM_PRIVN_TRANSACTION_ID = p.SHP_SRM_PRIVN_TRANSACTION_ID
{filtro_tipo}
GROUP BY b.Name, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE, b.INVOICE_IDENTIFIER_STATUS
ORDER BY b.Name ASC, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE, b.INVOICE_IDENTIFIER_STATUS
"""

QUERY_RV_TEMPLATE = """
WITH TopPeriods AS (
  SELECT DISTINCT JSON_EXTRACT_SCALAR(SHP_SRM_PRIVN_PERIOD, '$.name') AS Name
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  ORDER BY Name DESC
  LIMIT 5
),
LatestVersion AS (
  SELECT SHP_SRM_PRIVN_TRANSACTION_ID, MAX(SHP_SRM_PRIVN_VERSION_ID) AS MaxVersion
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  GROUP BY SHP_SRM_PRIVN_TRANSACTION_ID
),
Base AS (
  SELECT C.SHP_SRM_PRIVN_TRANSACTION_ID,
         JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name') AS Name,
         C.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
         C.SHP_SRM_PRIVN_PAYLOAD
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES C
  JOIN LatestVersion LV ON C.SHP_SRM_PRIVN_TRANSACTION_ID = LV.SHP_SRM_PRIVN_TRANSACTION_ID
                        AND C.SHP_SRM_PRIVN_VERSION_ID = LV.MaxVersion
  JOIN TopPeriods TP ON JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name') = TP.Name
  JOIN WHOWNER.BT_SHP_SRM_PROVIDERS PRV ON C.SHP_SRM_PRIVN_PROVIDER_ID = PRV.ID
  WHERE C.SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
    AND PRV.TYPE IN ('TAC-CNPJ', 'TAC-CPF', 'ETC', 'MEI')
),
PayloadRV AS (
  SELECT b.SHP_SRM_PRIVN_TRANSACTION_ID,
         b.Name,
         b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
         MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'external_route_id', JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS external_route_id,
         MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'vehicle_id',        JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS vehicle_id
  FROM Base b, UNNEST(JSON_EXTRACT_ARRAY(b.SHP_SRM_PRIVN_PAYLOAD, '$.payload')) AS p
  GROUP BY b.SHP_SRM_PRIVN_TRANSACTION_ID, b.Name, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE
)
SELECT Name,
       COUNT(DISTINCT external_route_id) AS qtd_rotas_distintas,
       COUNT(DISTINCT vehicle_id)        AS qtd_veiculos_distintos
FROM PayloadRV
{filtro_tipo_rv}
GROUP BY Name
ORDER BY Name ASC
"""

QUERY_MENSAL_TEMPLATE = """
WITH TopMonths AS (
  SELECT DISTINCT SUBSTR(JSON_EXTRACT_SCALAR(SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) AS Month
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  ORDER BY Month DESC
  LIMIT 5
),
LatestVersion AS (
  SELECT SHP_SRM_PRIVN_TRANSACTION_ID, MAX(SHP_SRM_PRIVN_VERSION_ID) AS MaxVersion
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  GROUP BY SHP_SRM_PRIVN_TRANSACTION_ID
),
TraceDedup AS (
  SELECT PREINVOICE_ID, INVOICE_IDENTIFIER_STATUS
  FROM (
    SELECT PREINVOICE_ID, INVOICE_IDENTIFIER_STATUS,
           ROW_NUMBER() OVER (PARTITION BY PREINVOICE_ID ORDER BY INVOICE_LAST_UPDATE DESC) AS rn
    FROM WHOWNER.BT_SRM_INVOICE_FILE_TRACEABILITY
  )
  WHERE rn = 1
),
Base AS (
  SELECT
    C.SHP_SRM_PRIVN_TRANSACTION_ID,
    SUBSTR(JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) AS Month,
    C.SHP_SRM_PRIVN_COST_AMT,
    C.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
    C.SHP_SRM_PRIVN_PRE_INVOICE_ID,
    C.SHP_SRM_PRIVN_PAYLOAD,
    PRV.TYPE AS PROVIDER_TYPE,
    CASE
      WHEN PRV.TYPE = 'TAC-CPF' AND T.INVOICE_IDENTIFIER_STATUS IS NULL THEN 'nao_emite_nfs'
      WHEN T.INVOICE_IDENTIFIER_STATUS IN ('sap_status_pago_realizado', 'sap_status_pagado') THEN 'sap_status_pago_realizado'
      ELSE T.INVOICE_IDENTIFIER_STATUS
    END AS INVOICE_IDENTIFIER_STATUS
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES C
  JOIN LatestVersion LV
    ON C.SHP_SRM_PRIVN_TRANSACTION_ID = LV.SHP_SRM_PRIVN_TRANSACTION_ID
   AND C.SHP_SRM_PRIVN_VERSION_ID = LV.MaxVersion
  JOIN TopMonths TM
    ON SUBSTR(JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) = TM.Month
  JOIN WHOWNER.BT_SHP_SRM_PROVIDERS PRV
    ON C.SHP_SRM_PRIVN_PROVIDER_ID = PRV.ID
  LEFT JOIN TraceDedup T
    ON CAST(C.SHP_SRM_PRIVN_PRE_INVOICE_ID AS STRING) = T.PREINVOICE_ID
  WHERE C.SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
    AND PRV.TYPE IN ('TAC-CNPJ', 'TAC-CPF', 'ETC', 'MEI')
),
PayloadKeyVals AS (
  SELECT
    b.SHP_SRM_PRIVN_TRANSACTION_ID,
    MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'external_route_id', JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS external_route_id
  FROM Base b,
       UNNEST(JSON_EXTRACT_ARRAY(b.SHP_SRM_PRIVN_PAYLOAD, '$.payload')) AS p
  GROUP BY b.SHP_SRM_PRIVN_TRANSACTION_ID
)
SELECT
  b.Month,
  b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
  b.INVOICE_IDENTIFIER_STATUS,
  COUNT(DISTINCT p.external_route_id)            AS qtd_rotas_distintas,
  COUNT(DISTINCT b.SHP_SRM_PRIVN_PRE_INVOICE_ID) AS qtd_pre_invoice_id_distintos,
  SUM(b.SHP_SRM_PRIVN_COST_AMT)                  AS SHP_SRM_PRIVN_COST_AMT
FROM Base b
LEFT JOIN PayloadKeyVals p
  ON b.SHP_SRM_PRIVN_TRANSACTION_ID = p.SHP_SRM_PRIVN_TRANSACTION_ID
{filtro_tipo}
GROUP BY b.Month, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE, b.INVOICE_IDENTIFIER_STATUS
ORDER BY b.Month ASC, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE, b.INVOICE_IDENTIFIER_STATUS
"""

QUERY_RV_MENSAL_TEMPLATE = """
WITH TopMonths AS (
  SELECT DISTINCT SUBSTR(JSON_EXTRACT_SCALAR(SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) AS Month
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  ORDER BY Month DESC
  LIMIT 5
),
LatestVersion AS (
  SELECT SHP_SRM_PRIVN_TRANSACTION_ID, MAX(SHP_SRM_PRIVN_VERSION_ID) AS MaxVersion
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES
  WHERE SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
  GROUP BY SHP_SRM_PRIVN_TRANSACTION_ID
),
Base AS (
  SELECT C.SHP_SRM_PRIVN_TRANSACTION_ID,
         SUBSTR(JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) AS Month,
         C.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
         C.SHP_SRM_PRIVN_PAYLOAD
  FROM WHOWNER.BT_SRM_PRE_INVOICE_ASSIGNED_EVENT_COST_ENTITIES C
  JOIN LatestVersion LV ON C.SHP_SRM_PRIVN_TRANSACTION_ID = LV.SHP_SRM_PRIVN_TRANSACTION_ID
                        AND C.SHP_SRM_PRIVN_VERSION_ID = LV.MaxVersion
  JOIN TopMonths TM ON SUBSTR(JSON_EXTRACT_SCALAR(C.SHP_SRM_PRIVN_PERIOD, '$.name'), 1, 6) = TM.Month
  JOIN WHOWNER.BT_SHP_SRM_PROVIDERS PRV ON C.SHP_SRM_PRIVN_PROVIDER_ID = PRV.ID
  WHERE C.SHP_SRM_PRIVN_EVENT_STATUS = 'PROCESSED'
    AND PRV.TYPE IN ('TAC-CNPJ', 'TAC-CPF', 'ETC', 'MEI')
),
PayloadRV AS (
  SELECT b.SHP_SRM_PRIVN_TRANSACTION_ID,
         b.Month,
         b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE,
         MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'external_route_id', JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS external_route_id,
         MAX(IF(JSON_EXTRACT_SCALAR(p, '$.key') = 'vehicle_id',        JSON_EXTRACT_SCALAR(p, '$.value'), NULL)) AS vehicle_id
  FROM Base b, UNNEST(JSON_EXTRACT_ARRAY(b.SHP_SRM_PRIVN_PAYLOAD, '$.payload')) AS p
  GROUP BY b.SHP_SRM_PRIVN_TRANSACTION_ID, b.Month, b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE
)
SELECT Month,
       COUNT(DISTINCT external_route_id) AS qtd_rotas_distintas,
       COUNT(DISTINCT vehicle_id)        AS qtd_veiculos_distintos
FROM PayloadRV
{filtro_tipo_rv}
GROUP BY Month
ORDER BY Month ASC
"""

# ============================================================
# Constantes de processamento
# ============================================================
OUTROS = [
    'accounting_document_update', 'accounting_error', 'accounting_sent_sap',
    'escalated_to_external_team', 'validate_files_fail',
    'validate_provision_success', 'preconciliation_ok_invoice'
]

ORDEM_STATUS = [
    'sap_status_pago_realizado',
    'sap_status_liberado_para_el_pago',
    'sap_status_en_proceso_de_pago',
    'nao_emite_nfs',
    'preconciliation_fail_invoice',
    'NULL',
    'Outros',
    'TOTAL',
]

def normalizar_status(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return 'NULL'
    if s in OUTROS:
        return 'Outros'
    return s

def build_pivot(df, periodos, metric):
    agg = df.groupby(['Name', 'status_norm'])[metric].sum().reset_index()
    piv = agg.pivot(index='status_norm', columns='Name', values=metric).fillna(0)
    for s in ORDEM_STATUS[:-1]:
        if s not in piv.index:
            piv.loc[s] = 0
    piv.loc['TOTAL'] = piv.sum()
    ordem_presente = [s for s in ORDEM_STATUS if s in piv.index or s == 'TOTAL']
    piv = piv.loc[ordem_presente]
    piv = piv[periodos]
    return piv

def build_pivot_mensal(df_m, periodos_m, metric):
    agg = df_m.groupby(['Month', 'status_norm'])[metric].sum().reset_index()
    piv = agg.pivot(index='status_norm', columns='Month', values=metric).fillna(0)
    for s in ORDEM_STATUS[:-1]:
        if s not in piv.index:
            piv.loc[s] = 0
    piv.loc['TOTAL'] = piv.sum()
    ordem_presente = [s for s in ORDEM_STATUS if s in piv.index or s == 'TOTAL']
    piv = piv.loc[ordem_presente]
    piv = piv[periodos_m]
    return piv

# ============================================================
# Impressao dos quadros no terminal
# ============================================================
SEP = ' | '
status_w = 40
col_w = 16
metrica_w = 25

def fmt_qtd(v):
    return '-' if v == 0 else f'{int(v):,}'.replace(',', '.')

def fmt_custo(v):
    if v == 0:
        return '-'
    s = f'R$ {v:,.0f}'
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')

def print_quadro(piv, fmt_fn, title, is_pct=False):
    print(f'\n{"="*85}')
    print(f'  {title}')
    print(f'{"="*85}')
    cols = list(piv.columns)
    header = f'{"Status":<{status_w}}' + SEP.join(f'{c:>{col_w}}' for c in cols)
    print(header)
    print('-'*len(header))
    for idx, row in piv.iterrows():
        if is_pct:
            vals = [('100%' if idx == 'TOTAL' else ('-' if row[c] == 0 else f'{row[c]:.1f}%')) for c in cols]
        else:
            vals = [fmt_fn(row[c]) for c in cols]
        label = f'** {idx} **' if idx == 'TOTAL' else idx
        print(f'{label:<{status_w}}' + SEP.join(f'{v:>{col_w}}' for v in vals))

def print_rv(vals_idx, periodos_rv, titulo, metrica_col, label_str):
    print(f'\n{"="*85}')
    print(f'  {titulo}')
    print(f'{"="*85}')
    h = f'{"Metrica":<{metrica_w}}' + SEP.join(f'{p:>{col_w}}' for p in periodos_rv)
    print(h)
    print('-'*len(h))
    c = SEP.join(
        f'{int(vals_idx.loc[p, metrica_col]):>{col_w},}'.replace(',', '.') if p in vals_idx.index else f'{"-":>{col_w}}'
        for p in periodos_rv
    )
    print(f'{label_str:<{metrica_w}}{c}')

# ============================================================
# Loop por tipo
# ============================================================
TIPOS = ['regular', 'complementary', 'ambos']
resultados = {}

for TIPO in TIPOS:
    print(f'\n{"#"*85}')
    print(f'# Processando: {TIPO.upper()}')
    print(f'{"#"*85}')

    if TIPO == 'ambos':
        filtro_tipo    = ''
        filtro_tipo_rv = ''
    elif TIPO == 'complementary':
        filtro_tipo    = "WHERE b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE = 'complementary'"
        filtro_tipo_rv = "WHERE SHP_SRM_PRIVN_PRE_INVOICE_TYPE = 'complementary'"
    else:
        filtro_tipo    = f"WHERE b.SHP_SRM_PRIVN_PRE_INVOICE_TYPE = '{TIPO}'"
        filtro_tipo_rv = f"WHERE SHP_SRM_PRIVN_PRE_INVOICE_TYPE = '{TIPO}'"

    query           = QUERY_TEMPLATE.format(filtro_tipo=filtro_tipo)
    query_rv        = QUERY_RV_TEMPLATE.format(filtro_tipo_rv=filtro_tipo_rv)
    query_mensal    = QUERY_MENSAL_TEMPLATE.format(filtro_tipo=filtro_tipo)
    query_rv_mensal = QUERY_RV_MENSAL_TEMPLATE.format(filtro_tipo_rv=filtro_tipo_rv)

    print('Executando query principal...')
    df = client.query(query).to_dataframe()
    print(f'Linhas: {len(df)}')

    print('Executando query rotas/veiculos...')
    df_rv = client.query(query_rv).to_dataframe()
    print(f'Linhas: {len(df_rv)}')

    print('Executando query mensal principal...')
    df_mensal = client.query(query_mensal).to_dataframe()
    print(f'Linhas: {len(df_mensal)}')

    print('Executando query mensal rotas/veiculos...')
    df_rv_mensal = client.query(query_rv_mensal).to_dataframe()
    print(f'Linhas: {len(df_rv_mensal)}')

    if df.empty:
        print(f'Sem dados para tipo "{TIPO}". Pulando...')
        resultados[TIPO] = {'sem_dados': True}
        continue

    # Pivots semanais
    df['status_norm'] = df['INVOICE_IDENTIFIER_STATUS'].apply(normalizar_status)
    periodos = sorted(df['Name'].unique())
    piv_qtd   = build_pivot(df, periodos, 'qtd_pre_invoice_id_distintos')
    piv_custo = build_pivot(df, periodos, 'SHP_SRM_PRIVN_COST_AMT')
    piv_pct   = piv_qtd.div(piv_qtd.loc['TOTAL']).mul(100).round(1).fillna(0)
    piv_pct.loc['TOTAL'] = 100.0

    # Pivots mensais
    df_mensal['status_norm'] = df_mensal['INVOICE_IDENTIFIER_STATUS'].apply(normalizar_status)
    periodos_mensal = sorted(df_mensal['Month'].unique())
    piv_qtd_m   = build_pivot_mensal(df_mensal, periodos_mensal, 'qtd_pre_invoice_id_distintos')
    piv_custo_m = build_pivot_mensal(df_mensal, periodos_mensal, 'SHP_SRM_PRIVN_COST_AMT')
    piv_pct_m   = piv_qtd_m.div(piv_qtd_m.loc['TOTAL']).mul(100).round(1).fillna(0)
    piv_pct_m.loc['TOTAL'] = 100.0

    # RV semanais/mensais
    df_rv_s    = df_rv.sort_values('Name')
    vals_idx   = df_rv_s.set_index('Name')
    df_rv_ms   = df_rv_mensal.sort_values('Month')
    vals_idx_m = df_rv_ms.set_index('Month')
    periodos_rv   = sorted(vals_idx.index.tolist())
    periodos_rv_m = sorted(vals_idx_m.index.tolist())

    # Imprimir quadros no terminal
    print_quadro(piv_qtd,   fmt_qtd,   f'QUADRO 1 -- Quantidade de Pre-invoices ({TIPO})')
    print_quadro(piv_pct,   None,      f'QUADRO 2 -- % sobre total de pre-invoices do periodo ({TIPO})', is_pct=True)
    print_quadro(piv_custo, fmt_custo, f'QUADRO 3 -- Custo total em R$ ({TIPO})')
    print_rv(vals_idx,   periodos_rv,   f'QUADRO 4  -- Rotas distintas por periodo ({TIPO})',   'qtd_rotas_distintas',    'Rotas distintas')
    print_rv(vals_idx,   periodos_rv,   f'QUADRO 5  -- Veiculos distintos por periodo ({TIPO})', 'qtd_veiculos_distintos', 'Veiculos distintos')
    print_quadro(piv_qtd_m,   fmt_qtd,   f'QUADRO 6  -- Quantidade de Pre-invoices por MES ({TIPO})')
    print_quadro(piv_pct_m,   None,      f'QUADRO 7  -- % sobre total de pre-invoices por MES ({TIPO})', is_pct=True)
    print_quadro(piv_custo_m, fmt_custo, f'QUADRO 8  -- Custo total em R$ por MES ({TIPO})')
    print_rv(vals_idx_m, periodos_rv_m, f'QUADRO 9  -- Rotas distintas por MES ({TIPO})',        'qtd_rotas_distintas',    'Rotas distintas')
    print_rv(vals_idx_m, periodos_rv_m, f'QUADRO 10 -- Veiculos distintos por MES ({TIPO})',     'qtd_veiculos_distintos', 'Veiculos distintos')

    resultados[TIPO] = {
        'periodos':    periodos,
        'periodos_m':  periodos_mensal,
        'piv_qtd':     piv_qtd,
        'piv_pct':     piv_pct,
        'piv_custo':   piv_custo,
        'df_rv':       vals_idx,
        'periodos_rv':   periodos_rv,
        'piv_qtd_m':   piv_qtd_m,
        'piv_pct_m':   piv_pct_m,
        'piv_custo_m': piv_custo_m,
        'df_rv_m':     vals_idx_m,
        'periodos_rv_m': periodos_rv_m,
    }

# ============================================================
# Export Excel com openpyxl
# ============================================================
FILL_HEADER_ABA = PatternFill('solid', fgColor='002060')  # azul escuro MeLi
FILL_TITULO_Q   = PatternFill('solid', fgColor='1F4E79')  # azul medio
FILL_HEADER_COL = PatternFill('solid', fgColor='D6E4F0')  # azul muito claro
FILL_TOTAL      = PatternFill('solid', fgColor='FFF2CC')  # amarelo claro
FONT_WHITE_BOLD = Font(bold=True, color='FFFFFF')
FONT_BOLD       = Font(bold=True)
ALIGN_CENTER    = Alignment(horizontal='center')
ALIGN_LEFT      = Alignment(horizontal='left')


def write_pivot_to_sheet(ws, piv, titulo, row_start, is_pct=False, is_custo=False):
    cols = list(piv.columns)
    n_cols = len(cols)

    # Titulo do quadro
    cell = ws.cell(row=row_start, column=1, value=titulo)
    cell.font = FONT_WHITE_BOLD
    cell.fill = FILL_TITULO_Q
    cell.alignment = ALIGN_CENTER
    ws.merge_cells(start_row=row_start, start_column=1,
                   end_row=row_start, end_column=1 + n_cols)
    row_start += 1

    # Header de colunas
    hcell = ws.cell(row=row_start, column=1, value='Status')
    hcell.fill = FILL_HEADER_COL
    hcell.font = FONT_BOLD
    hcell.alignment = ALIGN_LEFT
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(row=row_start, column=j, value=c)
        cell.fill = FILL_HEADER_COL
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
    row_start += 1

    # Linhas de dados
    for idx, row in piv.iterrows():
        is_total = (idx == 'TOTAL')
        ws.cell(row=row_start, column=1, value=idx)
        if is_total:
            ws.cell(row=row_start, column=1).font = FONT_BOLD
            ws.cell(row=row_start, column=1).fill = FILL_TOTAL
        for j, c in enumerate(cols, start=2):
            v = row[c]
            if is_pct:
                safe_v = 0 if pd.isna(v) else v / 100
                cell = ws.cell(row=row_start, column=j, value=safe_v)
                cell.number_format = '0.0%'
            elif is_custo:
                cell = ws.cell(row=row_start, column=j, value=float(v))
                cell.number_format = '"R$ "#,##0'
            else:
                cell = ws.cell(row=row_start, column=j, value=int(v))
                cell.number_format = '#,##0'
            cell.alignment = ALIGN_CENTER
            if is_total:
                cell.font = FONT_BOLD
                cell.fill = FILL_TOTAL
        row_start += 1

    return row_start + 1  # +1 linha vazia


def write_rv_to_sheet(ws, vals_idx, periodos, titulo, metrica_col, label_str, row_start):
    n_cols = len(periodos)

    # Titulo do quadro
    cell = ws.cell(row=row_start, column=1, value=titulo)
    cell.font = FONT_WHITE_BOLD
    cell.fill = FILL_TITULO_Q
    cell.alignment = ALIGN_CENTER
    ws.merge_cells(start_row=row_start, start_column=1,
                   end_row=row_start, end_column=1 + n_cols)
    row_start += 1

    # Header
    hcell = ws.cell(row=row_start, column=1, value='Metrica')
    hcell.fill = FILL_HEADER_COL
    hcell.font = FONT_BOLD
    hcell.alignment = ALIGN_LEFT
    for j, p in enumerate(periodos, start=2):
        cell = ws.cell(row=row_start, column=j, value=p)
        cell.fill = FILL_HEADER_COL
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
    row_start += 1

    # Linha de dados
    ws.cell(row=row_start, column=1, value=label_str)
    for j, p in enumerate(periodos, start=2):
        v = int(vals_idx.loc[p, metrica_col]) if p in vals_idx.index else 0
        cell = ws.cell(row=row_start, column=j, value=v)
        cell.number_format = '#,##0'
        cell.alignment = ALIGN_CENTER
    row_start += 1

    return row_start + 1  # +1 linha vazia


def exportar_excel(resultados, caminho):
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrao

    agora = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    NOMES_ABAS = {'regular': 'Regular', 'complementary': 'Complementary', 'ambos': 'Ambos'}

    for tipo in ['regular', 'complementary', 'ambos']:
        r  = resultados[tipo]
        ws = wb.create_sheet(title=NOMES_ABAS[tipo])

        if r.get('sem_dados'):
            cell = ws.cell(row=1, column=1, value=f'Sem dados para pre-invoices do tipo "{tipo}".')
            cell.font = FONT_BOLD
            ws.column_dimensions['A'].width = 50
            continue

        periodos_s    = r['periodos']
        periodos_m    = r['periodos_m']
        periodos_rv_s = r['periodos_rv']
        periodos_rv_m = r['periodos_rv_m']
        n_cols_max = max(len(periodos_s), len(periodos_m),
                         len(periodos_rv_s), len(periodos_rv_m)) + 1

        # Linha 1: titulo da aba
        titulo_aba = f'Faturamento — {NOMES_ABAS[tipo]} | Extracao: {agora}'
        cell = ws.cell(row=1, column=1, value=titulo_aba)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_HEADER_ABA
        cell.alignment = ALIGN_CENTER
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols_max)

        row = 3  # linha 2 vazia

        # ---- SEMANAL ----
        row = write_pivot_to_sheet(ws, r['piv_qtd'],
            'QUADRO 1 — Quantidade de Pre-invoices (semanal)', row)
        row = write_pivot_to_sheet(ws, r['piv_pct'],
            'QUADRO 2 — % sobre total de pre-invoices (semanal)', row, is_pct=True)
        row = write_pivot_to_sheet(ws, r['piv_custo'],
            'QUADRO 3 — Custo total em R$ (semanal)', row, is_custo=True)
        row = write_rv_to_sheet(ws, r['df_rv'], periodos_rv_s,
            'QUADRO 4 — Rotas distintas por periodo (semanal)',
            'qtd_rotas_distintas', 'Rotas distintas', row)
        row = write_rv_to_sheet(ws, r['df_rv'], periodos_rv_s,
            'QUADRO 5 — Veiculos distintos por periodo (semanal)',
            'qtd_veiculos_distintos', 'Veiculos distintos', row)

        # ---- MENSAL ----
        row = write_pivot_to_sheet(ws, r['piv_qtd_m'],
            'QUADRO 6 — Quantidade de Pre-invoices (mensal)', row)
        row = write_pivot_to_sheet(ws, r['piv_pct_m'],
            'QUADRO 7 — % sobre total de pre-invoices (mensal)', row, is_pct=True)
        row = write_pivot_to_sheet(ws, r['piv_custo_m'],
            'QUADRO 8 — Custo total em R$ (mensal)', row, is_custo=True)
        row = write_rv_to_sheet(ws, r['df_rv_m'], periodos_rv_m,
            'QUADRO 9 — Rotas distintas por MES (mensal)',
            'qtd_rotas_distintas', 'Rotas distintas', row)
        row = write_rv_to_sheet(ws, r['df_rv_m'], periodos_rv_m,
            'QUADRO 10 — Veiculos distintos por MES (mensal)',
            'qtd_veiculos_distintos', 'Veiculos distintos', row)

        # Largura das colunas
        ws.column_dimensions['A'].width = 45
        for col_idx in range(2, n_cols_max + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(caminho)
    print(f'\nExcel salvo em: {caminho}')


# ============================================================
# Gerar e abrir arquivo
# ============================================================
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
caminho_excel = os.path.join(downloads, f'faturamento_{ts}.xlsx')

exportar_excel(resultados, caminho_excel)
os.startfile(caminho_excel)
print('\nConcluido.')
